# -*- coding: utf-8 -*-


import logging
import pandas
import openpyxl
import os


# create logger
module_logger = logging.getLogger('ddd_check')


def lire_classeur(path, sheet_name):

    # Lecture du fichier directement dans une data frame
    xl = pandas.ExcelFile(path)
    df = xl.parse(sheet_name)

    desc_matrice = df.describe()

    # Infos debug
    module_logger.info('lecture classeur --> ' + path)
    module_logger.info('lecture feuille --> ' + sheet_name)
    # module_logger.info('champs = ' + desc_matrice)

    liste_df_brute = df.values.tolist()
    for item in liste_df_brute:
        module_logger.debug(item)

    module_logger.info("********************************************************************************")

    return (df)
    
def lire_pkl(file_path):

    # Infos debug
    module_logger.info('lecture pkl --> ' + file_path)

    df = pandas.read_pickle(file_path)

    desc_matrice = df.describe()

    liste_df_brute = df.values.tolist()
    for item in liste_df_brute:
        module_logger.debug(item)

    module_logger.info("********************************************************************************")

    return (df)    


def generate_pkl_from_excel(ddd_xl_file_path, sheet_name, ddd_pkl_file_path):

    module_logger.info('generation fichier pandas pickle')
    module_logger.info('fichier source = ' + ddd_xl_file_path)
    module_logger.info('fichier dest = ' + ddd_pkl_file_path)
    
    df = lire_classeur(ddd_xl_file_path, sheet_name)
    
    
    df.to_pickle(ddd_pkl_file_path)  # where to save it, usually as a .pkl

'''
    Sauvegarde d'une liste dans un excel

'''
def sauvegarder_liste_in_excel_xslx(liste_xl, dest_filename, sheet_name):

    module_logger.info("********************************************************************************")

    module_logger.info('Sauvegarde --> ' + dest_filename)
    
    # Le fichier n'existe pas
    if (os.path.exists(dest_filename) == False):
        module_logger.info("fichier existe=NON")
        wb = openpyxl.Workbook()
        wb.create_sheet(title=sheet_name)
        ws = wb.get_sheet_by_name(sheet_name)
    # il existe
    else:
        module_logger.info("fichier existe=OUI")
        # Test si la feuille existe
        wb = openpyxl.load_workbook(dest_filename)
        list_sheets = wb.get_sheet_names()
        if not (sheet_name in list_sheets):
            module_logger.info("feuille = CREATION")
            wb.create_sheet(title=sheet_name)
        else:
            module_logger.info("feuille = EFACEMENT+CREATION")
            ws = wb.get_sheet_by_name(sheet_name)
            wb.remove_sheet(ws)
            wb.create_sheet(title=sheet_name)
            ws = wb.get_sheet_by_name(sheet_name)

    ws = wb.get_sheet_by_name(sheet_name)
    
    
    ws = wb.get_sheet_by_name(sheet_name)

    i_row = 1
    for record in liste_xl:
        i_col = 1
        module_logger.debug(record)
        for elt in record:
            ws.cell(column=i_col, row=i_row, value=str(elt))
            i_col += 1
        i_row += 1
    wb.save(filename = dest_filename)

    module_logger.info("********************************************************************************")
    