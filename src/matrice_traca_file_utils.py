# -*- coding: utf-8 -*-

import logging
from openpyxl import Workbook
from openpyxl import load_workbook
import pandas





# create logger
module_logger = logging.getLogger('matrice_traca')

'''
    Sauvegarde matrice brute de la dataframe de base dans un excel

    ATTENTION : cette fonction doit etre appellee en premier, pour creer un classeur
    sinon remplacer la ligne
    "wb = Workbook()"
    par
    "wb = load_workbook(dest_filename)"

'''
def sauvegarder_df_matrice_brute_in_excel_xslx(df_matrice_stbl, dest_filename, sheet_name):

    module_logger.info("********************************************************************************")

    module_logger.info('Sauvegarde --> ' + dest_filename)

    matrice_stbl = df_matrice_stbl.values.tolist()

    wb = Workbook()
    wb.create_sheet(title=sheet_name)
    ws = wb.get_sheet_by_name(sheet_name)

    module_logger.info("creation des colonnes")

    # creation des titres de colone
    # ft = Font(bold=True)
    # _.font = ft
    ws.cell(column=1, row=1, value="Exigence STBL")
    ws.cell(column=2, row=1, value="Version exigence")
    ws.cell(column=3, row=1, value="Exigence SF")
    ws.cell(column=4, row=1, value="Perimetre")
    ws.cell(column=5, row=1, value="UC")
    ws.cell(column=6, row=1, value="Titre")

    indice_row = 2
    for record in matrice_stbl:
        module_logger.debug(record)
        ws.cell(column=1, row=indice_row, value=str(record[0]))
        ws.cell(column=2, row=indice_row, value=str(record[1]))
        ws.cell(column=3, row=indice_row, value=str(record[2]))
        ws.cell(column=4, row=indice_row, value=str(record[3]))
        ws.cell(column=5, row=indice_row, value=str(record[4]))
        ws.cell(column=6, row=indice_row, value=str(record[5]))
        indice_row +=1

    wb.save(filename = dest_filename)

    module_logger.info("********************************************************************************")


'''
    Sauvegarde une dataframe python pandas dans un excel

'''
def sauvegarder_df_in_excel_xslx(data_frame, path, sheet):

    module_logger.info('Sauvegarde dataframe --> chemin = ' + path)
    module_logger.info('Sauvegarde dataframe --> sheet = ' + sheet)

    book = load_workbook(path)
    writer = pandas.ExcelWriter(path, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    # Convert the dataframe to an XlsxWriter Excel object.
    data_frame.to_excel(writer, sheet_name=sheet)

    # Close the Pandas Excel writer and output the Excel file.
    writer.save()


'''
    Sauvegarde matrice brute de la dataframe de base dans un excel

'''
def sauvegarder_liste_in_excel_xslx(liste_xl, dest_filename, sheet_name):

    module_logger.info("********************************************************************************")

    module_logger.info('Sauvegarde --> ' + dest_filename)

    wb = load_workbook(dest_filename)
    wb.create_sheet(title=sheet_name)
    ws = wb.get_sheet_by_name(sheet_name)

    i_row = 1
    for record in liste_xl:
        i_col = 1
        module_logger.debug(record)
        for elt in record:
            ws.cell(column=i_col, row=i_row, value=str(elt))
            i_col += 1
        i_row += 1
    wb.save(filename = dest_filename)

    module_logger.info("********************************************************************************")

